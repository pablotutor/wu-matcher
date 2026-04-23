"""
api.py
------
FastAPI backend para WU Matcher.

Endpoints:
  POST /upload-syllabi  – procesa PDFs y retorna resultado completo
  GET  /search-wu       – búsqueda de asignaturas WU por nombre o código
  GET  /health

Ejecutar:
  uvicorn app.api:app --reload --port 8000
"""

from __future__ import annotations

import asyncio
import datetime
import json
import logging
import re
import tempfile
import uuid
from pathlib import Path
from typing import List
import sys

log = logging.getLogger(__name__)

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import BackgroundTasks, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from scraper.parse_my_syllabus import parse_pdf
from pipeline.retrieval import HybridRetriever
from rag.generator import (
    generate_justification,
    _extract_section_text,
    _course_entry_by_code,
    _get_client as _get_llm_client,
    MODEL_NAME as LLM_MODEL,
)

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(
    title="WU Matcher API",
    description="Encuentra equivalencias entre asignaturas UAM y WU Vienna",
    version="4.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Áreas temáticas canónicas — 12 etiquetas, compartidas entre
# clasificación de optativas UAM y clasificación de cursos WU.
# ---------------------------------------------------------------------------

AREAS_VALID = {
    "MARKETING", "FINANZAS", "DATA_SCIENCE", "GESTIÓN",
    "ESTRATEGIA", "DERECHO", "TECNOLOGÍA", "SOSTENIBILIDAD",
    "RECURSOS_HUMANOS", "ENTREPRENEURSHIP", "IA", "PRODUCT",
}

_AREAS_LIST_STR = (
    "MARKETING, FINANZAS, DATA_SCIENCE, GESTIÓN, ESTRATEGIA, DERECHO, "
    "TECNOLOGÍA, SOSTENIBILIDAD, RECURSOS_HUMANOS, ENTREPRENEURSHIP, IA, PRODUCT"
)

_CLASSIFY_PROMPT_TEMPLATE = """\
Clasifica esta asignatura SOLO en las categorías PRINCIPALES y DIRECTAS (máximo 3-4).

Categorías disponibles (ÚNICAS):
{areas}

REGLAS:
- Solo EXPLÍCITAMENTE mencionadas o claramente implícitas
- Si es "IA aplicada a X" → incluye AMBAS: IA + X
- No tangenciales. Si duda, elige la más específica
- Máximo 4 categorías

Devuelve SOLO nombres separados por comas.

Asignatura: {name}

Contenidos:
{contents}"""


def _classify_areas(name: str, contents: str) -> list[str]:
    """Clasifica una asignatura UAM en las 12 áreas canónicas via LLM."""
    client = _get_llm_client()
    prompt = _CLASSIFY_PROMPT_TEMPLATE.format(
        areas=_AREAS_LIST_STR,
        name=name,
        contents=(contents or "")[:800],
    )
    response = client.chat(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
    )
    raw   = response.message.content.strip()
    areas = [a.strip().upper() for a in raw.replace("\n", ",").split(",")]
    valid = [a for a in areas if a in AREAS_VALID]
    return valid[:4] if valid else ["GESTIÓN"]


# ---------------------------------------------------------------------------
# Pydantic models para /match-optatives
# ---------------------------------------------------------------------------

class OptativeItem(BaseModel):
    code: str
    name: str
    areas_tematicas: list[str] = []


class MatchOptativesRequest(BaseModel):
    optatives: list[OptativeItem]
    user_interests: str


# ---------------------------------------------------------------------------
# In-memory job store for async processing
# ---------------------------------------------------------------------------

job_results: dict[str, dict] = {}

# ---------------------------------------------------------------------------
# WU courses classified index (code → {code, name, areas})
# ---------------------------------------------------------------------------

_WU_CLASSIFIED_PATH = _PROJECT_ROOT / "data" / "processed" / "wu_courses_classified.json"

def _load_wu_classified() -> dict[str, dict]:
    """Carga wu_courses_classified.json y devuelve un dict {code: course}."""
    if not _WU_CLASSIFIED_PATH.exists():
        import logging as _log
        _log.getLogger(__name__).warning(
            "wu_courses_classified.json no encontrado en %s. "
            "Ejecuta: python scripts/classify_wu_courses.py",
            _WU_CLASSIFIED_PATH,
        )
        return {}
    data = json.loads(_WU_CLASSIFIED_PATH.read_text(encoding="utf-8"))
    courses = data.get("courses", data) if isinstance(data, dict) else data
    return {c["code"]: c for c in courses if "code" in c}

wu_by_code: dict[str, dict] = _load_wu_classified()

# ---------------------------------------------------------------------------
# Lazy singleton retriever (se inicializa una sola vez al primer uso)
# ---------------------------------------------------------------------------

_retriever: HybridRetriever | None = None


def _get_retriever() -> HybridRetriever:
    global _retriever
    if _retriever is None:
        _retriever = HybridRetriever()
    return _retriever


# ---------------------------------------------------------------------------
# Pipeline: procesa un PDF y retorna el dict de resultado
# ---------------------------------------------------------------------------

def _process_one(pdf_bytes: bytes, filename: str) -> dict:
    """
    Orquesta el pipeline completo para un PDF:
      1. parse_pdf         → extrae code, name, credits, contents
      2. process_query_course → top 5 matches WU (hybrid retrieval)
      3. generate_justification → match_percentage, recommendation, topics
    """
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = Path(tmp.name)

    try:
        query_course = parse_pdf(tmp_path)
    finally:
        tmp_path.unlink(missing_ok=True)

    if query_course.get("parse_error"):
        raise ValueError(f"No se pudo parsear '{filename}': ¿es un plan docente UAM?")

    query_course.setdefault("source", "uam")

    # Fallback: si el regex no extrajo contenidos, usar el nombre como query mínima
    if not query_course.get("contents", "").strip():
        log.warning("[_process_one] '%s' sin contenidos extraídos — usando nombre como fallback.", filename)
        query_course["contents"] = query_course.get("name", filename)

    # ── Retrieval ─────────────────────────────────────────────────────────
    retriever = _get_retriever()
    top5_raw = retriever.process_query_course(query_course, top_courses=5)

    # ── Enriquecer con contents/outcomes para el generador ────────────────
    chunks_path = _PROJECT_ROOT / "data" / "processed" / "chunks.json"
    chunks_data = json.loads(chunks_path.read_text(encoding="utf-8"))

    top5_enriched = []
    for match in top5_raw:
        code = match.get("code", "")
        entry = _course_entry_by_code(chunks_data, code)
        top5_enriched.append({
            **match,
            "contents":          _extract_section_text(entry, "contents")          if entry else "",
            "learning_outcomes": _extract_section_text(entry, "learning_outcomes") if entry else "",
        })

    # ── Generación (LLM) ──────────────────────────────────────────────────
    justification = generate_justification(query_course, top5_enriched)
    llm_by_code: dict[str, dict] = {
        m.get("code", ""): m for m in justification.get("matches", [])
    }

    # ── Combinar retrieval + LLM ──────────────────────────────────────────
    matches = []
    for r in top5_raw:
        code = r.get("code", "")
        llm  = llm_by_code.get(code, {})
        matches.append({
            "rank":               r["rank"],
            "code":               code,
            "name":               r.get("name", ""),
            "credits":            r.get("credits", ""),
            "type":               r.get("type", ""),
            "match_percentage":   llm.get("match_percentage"),
            "recommendation":     llm.get("recommendation", ""),
            "overlapping_topics": llm.get("overlapping_topics", []),
            "missing_in_wu":      llm.get("missing_in_wu", []),
            "extra_in_wu":        llm.get("extra_in_wu", []),
            "schedule":           r.get("schedule", []),
        })

    return {
        "code":    query_course.get("code", ""),
        "name":    query_course.get("name", ""),
        "credits": query_course.get("credits", ""),
        "matches": matches,
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.post("/upload-syllabi")
async def upload_syllabi(files: List[UploadFile] = File(...)):
    """
    Procesa uno o varios PDFs de guías docentes UAM.
    Espera a que todos terminen y retorna la respuesta completa.
    """
    if not files:
        raise HTTPException(status_code=400, detail="Se requiere al menos un fichero PDF.")

    files_data: list[tuple[bytes, str]] = []
    for upload in files:
        if not upload.filename or not upload.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = await upload.read()
        files_data.append((pdf_bytes, upload.filename))

    if not files_data:
        raise HTTPException(status_code=400, detail="No se encontraron PDFs válidos.")

    courses: list[dict] = []
    errors:  list[dict] = []

    for pdf_bytes, filename in files_data:
        try:
            # asyncio.to_thread evita bloquear el event loop durante CPU/IO intensivo
            course_result = await asyncio.to_thread(_process_one, pdf_bytes, filename)
            courses.append(course_result)
        except Exception as exc:
            errors.append({"file": filename, "error": str(exc)})

    return {
        "status":  "success",
        "total":   len(files_data),
        "courses": courses,
        "errors":  errors,
    }


async def _process_all_async(job_id: uuid.UUID, files_data: list[tuple[bytes, str]]) -> None:
    for pdf_bytes, filename in files_data:
        try:
            course_result = await asyncio.to_thread(_process_one, pdf_bytes, filename)
            job_results[job_id]["courses"].append(course_result)
        except Exception as exc:
            job_results[job_id]["errors"].append({"file": filename, "error": str(exc)})
    job_results[job_id]["status"] = "done"


@app.post("/upload-syllabi-async")
async def upload_syllabi_async(background_tasks: BackgroundTasks, files: List[UploadFile] = File(...)):
    """Inicia el procesamiento en background y retorna un job_id para polling."""
    if not files:
        raise HTTPException(status_code=400, detail="Se requiere al menos un fichero PDF.")

    files_data: list[tuple[bytes, str]] = []
    for upload in files:
        if not upload.filename or not upload.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = await upload.read()
        files_data.append((pdf_bytes, upload.filename))

    if not files_data:
        raise HTTPException(status_code=400, detail="No se encontraron PDFs válidos.")

    job_id = uuid.uuid4()
    job_results[job_id] = {"courses": [], "errors": [], "status": "processing", "total": len(files_data)}
    background_tasks.add_task(_process_all_async, job_id, files_data)
    return {"job_id": str(job_id)}


@app.get("/job/{job_id}")
def get_job(job_id: uuid.UUID):
    job = job_results.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job no encontrado.")
    courses = job["courses"]
    return {
        "status":    job["status"],
        "courses":   courses,
        "total":     job["total"],
        "completed": len(courses),
    }


@app.post("/upload-optatives")
async def upload_optatives(files: List[UploadFile] = File(...)):
    """
    Procesa PDFs de optativas UAM: extrae info via parse_pdf y clasifica
    en áreas temáticas via LLM. Guarda en data/my_courses/{code}_optative.json.
    """
    if not files:
        raise HTTPException(status_code=400, detail="Se requiere al menos un fichero PDF.")

    my_courses_dir = _PROJECT_ROOT / "data" / "my_courses"
    my_courses_dir.mkdir(parents=True, exist_ok=True)

    results: list[dict] = []
    errors:  list[dict] = []

    for upload in files:
        if not upload.filename or not upload.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = await upload.read()
        filename  = upload.filename

        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(pdf_bytes)
                tmp_path = Path(tmp.name)
            try:
                query_course = await asyncio.to_thread(parse_pdf, tmp_path)
            finally:
                tmp_path.unlink(missing_ok=True)

            if query_course.get("parse_error"):
                raise ValueError(f"No se pudo parsear '{filename}'")

            code     = query_course.get("code", "")
            name     = query_course.get("name", "")
            credits  = query_course.get("credits", "")
            contents = query_course.get("contents", "")

            areas = await asyncio.to_thread(_classify_areas, name, contents)

            optative_data = {
                "code": code, "name": name, "credits": credits,
                "contents": contents, "areas_tematicas": areas,
                "source": "uam_optativa",
            }
            json_path = my_courses_dir / f"{code}_optative.json"
            json_path.write_text(
                json.dumps(optative_data, ensure_ascii=False, indent=2), encoding="utf-8"
            )

            results.append({"code": code, "name": name, "credits": credits, "areas_tematicas": areas})
        except Exception as exc:
            errors.append({"file": filename, "error": str(exc)})

    return {"optatives": results, "errors": errors}



# ---------------------------------------------------------------------------
# Modelos para /generate-llm-report y /generate-learning-agreement
# ---------------------------------------------------------------------------

class SelectedMatch(BaseModel):
    uam_code: str
    uam_name: str
    uam_credits: str = ""
    wu_code: str
    wu_name: str
    wu_credits: str = ""
    wu_type: str = ""
    areas: list[str] = []
    afinidad: int = 0


class GenerateReportRequest(BaseModel):
    selected_matches: list[SelectedMatch]
    user_interests: str
    all_alternatives: dict[str, list[dict]] = {}


class GenerateLARequest(BaseModel):
    selected_matches: list[SelectedMatch]


# ---------------------------------------------------------------------------
# Helpers: extracción de temario para el Excel
# ---------------------------------------------------------------------------

def _get_wu_topics(code: str, chunks_data: list) -> list[str]:
    entry = next((c for c in chunks_data if c["code"] == code), None)
    if not entry:
        return []
    topics: list[str] = []
    for chunk in entry.get("chunks", []):
        if chunk.get("section") == "contents":
            text = re.sub(r"^\[[A-Z_]+\]\s+\S+\s+\|\s*", "", chunk["text"]).strip()
            for line in text.split("\n"):
                line = line.strip()
                if line:
                    topics.append(line)
    return topics[:30]


def _get_uam_topics(code: str) -> list[str]:
    path = _PROJECT_ROOT / "data" / "my_courses" / f"{code}_optative.json"
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    contents = data.get("contents", "")
    return [ln.strip() for ln in contents.split("\n") if ln.strip()][:30]


def _get_uam_url(code: str) -> str:
    path = _PROJECT_ROOT / "data" / "my_courses" / f"{code}_optative.json"
    if path.exists():
        data = json.loads(path.read_text(encoding="utf-8"))
        if url := data.get("url", ""):
            return url
    return (
        f"https://secretaria-virtual.uam.es/doa/consultaPublica/look"
        f"%5bconpub%5dMostrarPubGuiaDocAs?entradaPublica=true"
        f"&idiomaPais=es.ES&_anoAcademico=2025&_codAsignatura={code}"
    )


# ---------------------------------------------------------------------------
# Helper: formatea temario via LLM para el Excel (síncrono, para to_thread)
# ---------------------------------------------------------------------------

_FORMAT_SYLLABUS_PROMPT = """\
Tienes el contenido de la asignatura: {course_name}

Texto original:
{contents_text}

Extrae y presenta SOLO:

1. OBJETIVO(S) DEL CURSO:
   Una frase clara sobre qué aprenderá el estudiante

2. TEMAS PRINCIPALES:
   - Tema 1
   - Tema 2
   - Tema 3
   (máximo 8 temas, uno por línea)

NO incluyas:
- Metodología
- Evaluación
- Requisitos previos
- Textos descriptivos

Formato EXACTO:
OBJETIVO:
[una línea clara]

TEMAS:
- [tema]
- [tema]
- [tema]

Devuelve SOLO esto, sin preámbulos."""


def _raw_bullets(text: str) -> str:
    """Formatea texto en bruto como lista de bullets, sin LLM."""
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()][:30]
    return "\n".join(f"• {ln}" for ln in lines) if lines else "—"


def _format_syllabus_for_excel(contents_text: str, course_name: str, course_type: str) -> str:
    if not contents_text.strip():
        return "—"
    try:
        client = _get_llm_client()
        prompt = _FORMAT_SYLLABUS_PROMPT.format(
            course_name=course_name,
            contents_text=contents_text[:2000],
        )
        response = client.chat(
            model=LLM_MODEL,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.message.content.strip()
    except Exception as exc:
        log.warning("[format_syllabus] LLM falló para '%s' (%s) — usando texto en bruto.", course_name, exc)
        return _raw_bullets(contents_text)


# ---------------------------------------------------------------------------
# Helper: construye el xlsx del Learning Agreement en memoria
# ---------------------------------------------------------------------------

def _build_la_xlsx(
    selected_matches: list[SelectedMatch],
    formatted_texts: dict[str, tuple[str, str]],  # uam_code -> (wu_text, uam_text)
) -> bytes:
    chunks_path = _PROJECT_ROOT / "data" / "processed" / "chunks.json"
    chunks_data: list = json.loads(chunks_path.read_text(encoding="utf-8")) if chunks_path.exists() else []

    wb = openpyxl.Workbook()

    # ── Estilos compartidos ────────────────────────────────────────────────────
    bold       = Font(bold=True)
    bold_sm    = Font(bold=True, size=10)
    bold_white = Font(bold=True, size=10, color="FFFFFF")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left  = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_mid = Alignment(horizontal="center", vertical="center")
    dest_fill  = PatternFill("solid", fgColor="E2EFDA")   # verde WU
    orig_fill  = PatternFill("solid", fgColor="C6D9F1")   # azul UAM
    hdr_fill   = PatternFill("solid", fgColor="F2F2F2")   # gris filas cabecera

    _s = Side(style="thin")
    border = Border(left=_s, right=_s, top=_s, bottom=_s)

    def _apply(cell, *, font=None, alignment=None, fill=None, brd=True):
        if font:      cell.font      = font
        if alignment: cell.alignment = alignment
        if fill:      cell.fill      = fill
        if brd:       cell.border    = border

    def _border_range(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = border

    # ── Hoja 1: Asignaturas a convalidar ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Asignaturas a convalidar"

    # Fila 1: cabeceras de institución
    ws1.merge_cells("A1:D1")
    ws1.merge_cells("E1:J1")
    for cell, label, fill in [
        (ws1["A1"], "VIENA (WU)",    dest_fill),
        (ws1["E1"], "MADRID (UAM)", orig_fill),
    ]:
        cell.value = label
        _apply(cell, font=Font(bold=True, size=13), alignment=center, fill=fill)

    ws1.row_dimensions[1].height = 28

    # Fila 2: cabeceras de columna
    col_headers = [
        "CÓDIGO", "ECTS", "NOMBRE ASIGNATURA", "PROGRAMA / TEMARIO (WU)",
        "PROGRAMA / TEMARIO (UAM)", "NOMBRE ASIGNATURA", "CARÁCTER", "SEMESTRE", "ECTS", "CÓDIGO",
    ]
    for j, h in enumerate(col_headers, 1):
        c = ws1.cell(2, j, h)
        _apply(c, font=bold_sm, alignment=center, fill=dest_fill if j <= 4 else orig_fill)
    ws1.row_dimensions[2].height = 36

    current_row = 3

    for match in selected_matches:
        wu_url  = f"https://learn.wu.ac.at/vvz-old/25w/{match.wu_code}"
        uam_url = _get_uam_url(match.uam_code)
        ects    = match.wu_credits or "6"

        wu_text, uam_text = formatted_texts.get(match.uam_code, ("—", "—"))

        # ── Fila de cabecera de la convalidación ─────────────────────────────
        row_data = [
            (1, match.wu_code,             bold, center_mid),
            (2, ects,                      None, center_mid),
            (3, match.wu_name,             bold, wrap_left),
            (4, "Ver ficha WU →",          None, center_mid),
            (5, "Ver guía UAM →",          None, center_mid),
            (6, match.uam_name,            bold, wrap_left),
            (7, "OPT",                     None, center_mid),
            (8, "",                        None, center_mid),
            (9, match.uam_credits or ects, None, center_mid),
            (10, match.uam_code,           bold, center_mid),
        ]
        for col, val, fnt, aln in row_data:
            c = ws1.cell(current_row, col, val)
            _apply(c, font=fnt, alignment=aln, fill=hdr_fill)

        ws1.cell(current_row, 4).hyperlink = wu_url
        ws1.cell(current_row, 4).style = "Hyperlink"
        ws1.cell(current_row, 5).hyperlink = uam_url
        ws1.cell(current_row, 5).style = "Hyperlink"

        ws1.row_dimensions[current_row].height = 30
        current_row += 1

        # ── Fila de temario (LLM-formateado) ─────────────────────────────────
        for col in range(1, 11):
            ws1.cell(current_row, col).border = border
            ws1.cell(current_row, col).alignment = wrap_left

        _apply(ws1.cell(current_row, 4, wu_text),  alignment=wrap_left)
        _apply(ws1.cell(current_row, 5, uam_text), alignment=wrap_left)

        n_lines = max(wu_text.count("\n"), uam_text.count("\n"), 3) + 1
        ws1.row_dimensions[current_row].height = max(n_lines * 16, 80)
        current_row += 1

        # ── Fila separadora ───────────────────────────────────────────────────
        ws1.row_dimensions[current_row].height = 8
        current_row += 1

    # Anchos de columna
    for col, w in zip("ABCDEFGHIJ", [14, 5, 30, 50, 50, 30, 9, 9, 5, 14]):
        ws1.column_dimensions[col].width = w

    # ── Hoja 2: Propuesta Acuerdo de Estudios ─────────────────────────────────
    ws2 = wb.create_sheet("Propuesta Acuerdo de Estudios")

    ws2.merge_cells("A1:C1")
    ws2.merge_cells("D1:H1")
    for cell, label, fill in [
        (ws2["A1"], "VIENA (WU)",    dest_fill),
        (ws2["D1"], "MADRID (UAM)", orig_fill),
    ]:
        cell.value = label
        _apply(cell, font=Font(bold=True, size=13), alignment=center, fill=fill)
    ws2.row_dimensions[1].height = 28

    h2 = ["CÓDIGO ASIGNATURA", "ECTS", "NOMBRE DE LA ASIGNATURA",
          "NOMBRE DE LA ASIGNATURA", "CARÁCTER", "SEMESTRE", "ECTS", "CÓDIGO"]
    for j, h in enumerate(h2, 1):
        c = ws2.cell(2, j, h)
        _apply(c, font=bold_sm, alignment=center, fill=dest_fill if j <= 3 else orig_fill)
    ws2.row_dimensions[2].height = 36

    for i, m in enumerate(selected_matches, 3):
        ects = m.wu_credits or "6"
        row_vals = [m.wu_code, ects, m.wu_name, m.uam_name, "OPT", "", m.uam_credits or ects, m.uam_code]
        for j, val in enumerate(row_vals, 1):
            c = ws2.cell(i, j, val)
            _apply(c, alignment=wrap_left if j in (3, 4) else center_mid)
        ws2.row_dimensions[i].height = 22

    total_row = len(selected_matches) + 3
    total_ects = sum(int(m.wu_credits) if m.wu_credits and m.wu_credits.isdigit() else 6 for m in selected_matches)
    for col, val in [(1, "TOTAL ECTS"), (2, total_ects), (7, total_ects)]:
        c = ws2.cell(total_row, col, val)
        _apply(c, font=bold, alignment=center_mid, fill=hdr_fill)
    ws2.row_dimensions[total_row].height = 22

    for col, w in zip("ABCDEFGH", [16, 5, 40, 40, 10, 10, 5, 14]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helper: llm report detallado (llamada síncrona, usada con to_thread)
# ---------------------------------------------------------------------------

_DETAILED_REPORT_PROMPT = """\
Analiza la selección de convalidaciones de un estudiante de Erasmus.

Perfil del usuario (areas de interes): {user_interests}

CONVALIDACIONES SELECCIONADAS:
{selected_list}

ALTERNATIVAS DISPONIBLES NO SELECCIONADAS (top 10 por asignatura):
{alternatives_list}

Genera un REPORT EJECUTIVO en Markdown con esta estructura exacta:

## Evaluacion por asignatura

{per_course_sections}

## Analisis global

Parrafo de 150-200 palabras. Como la combinacion cubre el perfil. Que competencias quedan cubiertas y cuales presentan gaps.

## Recomendaciones

- Orden sugerido de cursada (con justificacion breve)
- Sinergia entre asignaturas
- Consideraciones practicas (carga, prereqs)
- Maximo 4 bullets

Sin emojis. En espanol. Tono profesional y educativo. 800-1000 palabras en total."""

_PER_COURSE_SECTION = """\
### {uam_name} -> {wu_name}

**Puntos fuertes:** [1-2 frases sobre el encaje con el perfil del usuario]
**Puntos debiles:** [1-2 frases sobre gaps respecto al perfil]
**Justificacion frente a alternativas:** [1 frase explicando por que es mejor que las descartadas]\
"""


def _run_llm_report(body: GenerateReportRequest) -> dict:
    selected_list = "\n".join(
        f"- UAM: {m.uam_name} ({m.uam_code}) -> WU: {m.wu_name} ({m.wu_code}) | Afinidad: {m.afinidad}%"
        for m in body.selected_matches
    )

    alt_blocks: list[str] = []
    for m in body.selected_matches:
        alts = body.all_alternatives.get(m.uam_code, [])
        if not alts:
            continue
        lines = [f"  {i+1}. {a.get('name', '')} ({a.get('code', '')}) | {a.get('afinidad', 0)}%"
                 for i, a in enumerate(alts) if a.get("code") != m.wu_code]
        if lines:
            alt_blocks.append(f"Para '{m.uam_name}':\n" + "\n".join(lines))
    alternatives_list = "\n\n".join(alt_blocks) if alt_blocks else "Sin alternativas registradas."

    per_course_sections = "\n\n".join(
        _PER_COURSE_SECTION.format(uam_name=m.uam_name, wu_name=m.wu_name)
        for m in body.selected_matches
    )

    prompt = _DETAILED_REPORT_PROMPT.format(
        user_interests=body.user_interests,
        selected_list=selected_list,
        alternatives_list=alternatives_list,
        per_course_sections=per_course_sections,
    )

    client = _get_llm_client()
    response = client.chat(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
    )
    return {
        "summary": response.message.content.strip(),
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    }


@app.post("/match-optatives")
async def match_optativas(body: MatchOptativesRequest):
    """
    Para cada optativa UAM:
    1. Filtro HARD por área: solo cursos WU cuyas áreas clasificadas solapen
       con las áreas de la optativa (wu_courses_classified.json).
    2. Ranking SOFT por intereses: cosine_similarity(user_interests_emb, wu_emb)
       usando embeddings pre-computados en ChromaDB.
    3. Retorna top 10 con syllabus_url.
    """
    retriever      = _get_retriever()
    total_wu       = len(retriever._ids)
    matched: list[dict] = []

    # Texto de intereses (se usa igual para todas las optativas)
    interests_text = body.user_interests.strip()
    log.info("[match-optatives] User interests: %s", interests_text)

    for opt in body.optatives:
        # ── 1. Filtro HARD por etiqueta de área ──────────────────────────────
        areas_set = set(opt.areas_tematicas)

        if areas_set and wu_by_code:
            filtered_codes = [
                code for code, wu in wu_by_code.items()
                if any(area in wu.get("areas", []) for area in areas_set)
                and code in retriever._ids          # solo cursos indexados
            ]
            log.info(
                "[match-optatives] [%s] %s → %d/%d WU match",
                opt.code,
                ", ".join(sorted(areas_set)),
                len(filtered_codes),
                total_wu,
            )
        else:
            # Sin clasificación disponible: usar todos los cursos indexados
            filtered_codes = list(retriever._ids)
            log.warning(
                "[match-optatives] [%s] Sin áreas / wu_by_code vacío — usando %d cursos sin filtrar",
                opt.code, total_wu,
            )

        if not filtered_codes:
            # Área muy específica sin matches: caer en top global por intereses
            filtered_codes = list(retriever._ids)
            log.warning(
                "[match-optatives] [%s] Filtro de área devolvió 0 cursos — usando todos",
                opt.code,
            )

        # ── 2. Ranking SOFT por embedding de intereses ────────────────────────
        ranked: list[tuple[str, float]] = await asyncio.to_thread(
            retriever.rank_by_interests, interests_text, filtered_codes
        )

        # ── 3. Debug top 3 ────────────────────────────────────────────────────
        for pos, (code, score) in enumerate(ranked[:3], 1):
            name = (wu_by_code.get(code) or {}).get("name") or code
            log.info("  #%d [%s] %s — score=%.4f", pos, code, name, score)

        # ── 4. Enriquecer top 10 con metadata ─────────────────────────────────
        wu_top10: list[dict] = []
        for rank, (code, sim) in enumerate(ranked[:10], 1):
            try:
                idx  = retriever._ids.index(code)
                meta = retriever._metas[idx]
            except ValueError:
                continue
            wu_top10.append({
                "rank":         rank,
                "code":         code,
                "name":         meta["name"],
                "credits":      meta["credits"],
                "type":         meta["type"],
                "schedule":     json.loads(meta.get("schedule", "[]")),
                "afinidad":     max(1, round(sim * 100)),
                "syllabus_url": f"https://learn.wu.ac.at/vvz-old/25w/{code}",
            })

        matched.append({
            "uam_code":   opt.code,
            "uam_name":   opt.name,
            "areas":      opt.areas_tematicas,
            "wu_matches": wu_top10,
        })

    return {"optatives_matched": matched}


@app.post("/generate-llm-report")
async def generate_llm_report(body: GenerateReportRequest):
    """Genera un report ejecutivo en Markdown para las convalidaciones seleccionadas."""
    if not body.selected_matches:
        raise HTTPException(status_code=400, detail="Se requiere al menos una convalidación seleccionada.")
    try:
        result = await asyncio.to_thread(_run_llm_report, body)
    except Exception as exc:
        log.error("[generate-llm-report] LLM failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"Error generando el reporte: {exc}")
    return result


@app.post("/generate-learning-agreement")
async def generate_learning_agreement(body: GenerateLARequest):
    """Genera el Excel de Learning Agreement con las convalidaciones seleccionadas."""
    if not body.selected_matches:
        raise HTTPException(status_code=400, detail="Se requiere al menos una convalidación.")
    try:
        chunks_path = _PROJECT_ROOT / "data" / "processed" / "chunks.json"
        chunks_data: list = json.loads(chunks_path.read_text(encoding="utf-8")) if chunks_path.exists() else []

        # Formatea los temarios secuencialmente para evitar rate limiting (429)
        formatted_texts: dict[str, tuple[str, str]] = {}
        for m in body.selected_matches:
            wu_topics  = _get_wu_topics(m.wu_code, chunks_data)
            uam_topics = _get_uam_topics(m.uam_code)
            wu_raw  = "\n".join(wu_topics)
            uam_raw = "\n".join(uam_topics)

            fmt_wu = await asyncio.to_thread(_format_syllabus_for_excel, wu_raw, m.wu_name, "Destino (WU Vienna)")
            await asyncio.sleep(2)
            fmt_uam = await asyncio.to_thread(_format_syllabus_for_excel, uam_raw, m.uam_name, "Origen (UAM Madrid)")
            await asyncio.sleep(2)

            formatted_texts[m.uam_code] = (fmt_wu, fmt_uam)

        xlsx_bytes = await asyncio.to_thread(_build_la_xlsx, body.selected_matches, formatted_texts)
    except Exception as exc:
        log.error("[generate-la] failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"Error generando el Excel: {exc}")
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Learning_Agreement.xlsx"},
    )


@app.get("/wu-course/{code}")
def get_wu_course(code: str):
    """Retorna detalles completos de una asignatura WU por código."""
    chunks_path = _PROJECT_ROOT / "data" / "processed" / "chunks.json"
    chunks_data: list = json.loads(chunks_path.read_text(encoding="utf-8"))

    entry = next((c for c in chunks_data if c["code"] == code), None)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Asignatura '{code}' no encontrada.")

    meta = entry.get("metadata", {})

    def get_section(section: str) -> str:
        parts = [
            re.sub(r"^\[[A-Z_]+\]\s+\S+\s+\|\s*", "", ch["text"]).strip()
            for ch in entry.get("chunks", [])
            if ch["section"] == section
        ]
        return "\n\n".join(parts)

    return {
        "code":              entry["code"],
        "name":              entry["name"],
        "credits":           meta.get("credits", ""),
        "type":              meta.get("type", ""),
        "contents":          get_section("contents"),
        "learning_outcomes": get_section("learning_outcomes"),
        "schedule":          meta.get("schedule", []),
        "url":               f"https://learn.wu.ac.at/vvz-old/25w/{entry['code']}",
    }


@app.get("/search-wu")
def search_wu(q: str = "", limit: int = 10):
    """Búsqueda de asignaturas WU por nombre o código. Sin query devuelve las primeras 20."""
    chunks_path = _PROJECT_ROOT / "data" / "processed" / "chunks.json"
    chunks_data: list = json.loads(chunks_path.read_text(encoding="utf-8"))

    if not q.strip():
        subset = chunks_data[:20]
    else:
        q_lower = q.strip().lower()
        matches_exact = [c for c in chunks_data if c["code"] == q.strip()]
        matches_rest  = [
            c for c in chunks_data
            if c not in matches_exact
            and (q_lower in c["name"].lower() or q_lower in c["code"].lower())
        ]
        subset = (matches_exact + matches_rest)[:limit]

    return [
        {
            "code":     c["code"],
            "name":     c["name"],
            "credits":  c["metadata"].get("credits", ""),
            "type":     c["metadata"].get("type", ""),
            "schedule": c["metadata"].get("schedule", []),
        }
        for c in subset
    ]


@app.get("/health")
def health():
    return {"status": "ok"}
